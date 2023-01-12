from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook			# https://openpyxl.readthedocs.io/en/latest/api/openpyxl.workbook.workbook.html#openpyxl.workbook.workbook.Workbook
from openpyxl.worksheet.worksheet import Worksheet		# https://openpyxl.readthedocs.io/en/latest/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet
from openpyxl.cell.cell import Cell 					# https://openpyxl.readthedocs.io/en/latest/api/openpyxl.cell.cell.html#openpyxl.cell.cell.Cell

path = "C:/Users/Dalton/Downloads/Fluidigm Operating Model.xlsx"


class JsxElement:
	tag = ''
	def __init__(self, text='', children=[]):
		self.text = text
		self._children = children
	
	def _str(self):
		el = f"<{self.tag}>{self.text}"
		if self._children:
			el += '\n'+'\n'.join([str(c) for c in self._children])+'\n'
		el += f"</{self.tag}>"
		return el
	
	def __str__(self):
		return self._str()
	
	@property
	def tag(self):
		return self.tag
	
	@property
	def children(self):
		return self.children
	
	def add_child(self, child):
		self._children.append(child)


class JsxWorkbook(JsxElement):
	tag = 'Workbook'

class JsxWorksheet(JsxElement):
	tag = 'Worksheet'

class JsxCell(JsxElement):
	tag = 'Cell'





class RenderableWorkbook(Workbook, JsxWorkbook):
	def __init__(self, path):
		_wb = load_workbook(path)
		self.__dict__ = _wb.__dict__
		self._sheets = [RenderableWorksheet(s) for s in self._sheets]
		super(JsxWorkbook, self).__init__()

class RenderableWorksheet(Worksheet, JsxWorksheet):
	def __init__(self, ws):
		self.__dict__ = ws.__dict__
		self._cells = {k: RenderableCell(v) for k,v in self._cells.items()}
		super(JsxWorksheet, self).__init__()

class RenderableCell(Cell, JsxCell):
	__slots__ = '__dict__'
	def __init__(self, c):
		self.row = c.row
		self.column = c.column
		self._value = c._value
		self.data_type = c.data_type
		self.parent = c.parent
		self._hyperlink = c._hyperlink
		self._comment = c._comment
		super(JsxCell, self).__init__()



wb = RenderableWorkbook(path=path)

# print(wb['Operating Model'].values)

for row in wb['Operating Model']:
	# print(row)
	for c in row:
		print(c)

